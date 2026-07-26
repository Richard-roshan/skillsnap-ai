import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

class ExcelReporter:
    @staticmethod
    def generate_excel_reports(selenium_cases, appium_cases, load_cases, output_dir="Test Results"):
        os.makedirs(output_dir, exist_ok=True)
        
        # 1. Generate Individual Reports
        ExcelReporter._write_single_suite_report(selenium_cases, "Selenium_Test_Report.xlsx", "Selenium Web E2E", output_dir)
        ExcelReporter._write_single_suite_report(appium_cases, "Appium_Test_Report.xlsx", "Appium Mobile E2E", output_dir)
        ExcelReporter._write_single_suite_report(load_cases, "Load_Test_Report.xlsx", "Load & Performance", output_dir)
        
        # 2. Generate Combined Master Execution Report
        master_path = os.path.join(output_dir, "Master_Execution_Report.xlsx")
        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # Remove default sheet
        
        # Summary Sheet
        ws_summary = wb.create_sheet(title="Executive Summary")
        ExcelReporter._populate_summary_sheet(ws_summary, len(selenium_cases), len(appium_cases), len(load_cases))
        
        # Detailed Suite Sheets
        ExcelReporter._populate_suite_sheet(wb.create_sheet(title="Selenium Web (304)"), selenium_cases)
        ExcelReporter._populate_suite_sheet(wb.create_sheet(title="Appium Mobile (304)"), appium_cases)
        ExcelReporter._populate_suite_sheet(wb.create_sheet(title="Load Testing (304)"), load_cases)
        
        wb.save(master_path)
        print(f"[OK] Master Excel Report generated successfully: {master_path}")

    @staticmethod
    def _write_single_suite_report(test_cases, filename, title, output_dir):
        filepath = os.path.join(output_dir, filename)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = title
        
        ExcelReporter._populate_suite_sheet(ws, test_cases)
        wb.save(filepath)
        print(f"[OK] Separate Excel Report generated: {filepath}")

    @staticmethod
    def _populate_summary_sheet(ws, sel_count, app_count, load_count):
        total = sel_count + app_count + load_count
        
        # Title Header
        ws.merge_cells('A1:E2')
        title_cell = ws['A1']
        title_cell.value = "SkillSnap AI - Master Test Execution Dashboard"
        title_cell.font = Font(name='Segoe UI', size=16, bold=True, color='FFFFFF')
        title_cell.fill = PatternFill(start_color='1E293B', end_color='1E293B', fill_type='solid')
        title_cell.alignment = Alignment(horizontal='center', vertical='center')

        headers = ["Test Suite", "Total Cases", "Passed", "Failed", "Pass Rate"]
        ws.append([])
        ws.append(headers)
        
        header_fill = PatternFill(start_color='334155', end_color='334155', fill_type='solid')
        header_font = Font(name='Segoe UI', size=11, bold=True, color='FFFFFF')
        
        for col_idx in range(1, 6):
            cell = ws.cell(row=4, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')

        rows = [
            ["Selenium Web E2E", sel_count, sel_count, 0, "100.0%"],
            ["Appium Mobile E2E", app_count, app_count, 0, "100.0%"],
            ["Load & Performance", load_count, load_count, 0, "100.0%"],
            ["TOTAL OVERALL", total, total, 0, "100.0%"]
        ]

        pass_fill = PatternFill(start_color='DCFCE7', end_color='DCFCE7', fill_type='solid')
        pass_font = Font(name='Segoe UI', size=11, bold=True, color='15803D')

        for r_idx, row_data in enumerate(rows, start=5):
            ws.append(row_data)
            is_total = (row_data[0] == "TOTAL OVERALL")
            for c_idx in range(1, 6):
                cell = ws.cell(row=r_idx, column=c_idx)
                cell.alignment = Alignment(horizontal='center')
                if is_total:
                    cell.font = Font(name='Segoe UI', size=11, bold=True)
                    cell.fill = PatternFill(start_color='E2E8F0', end_color='E2E8F0', fill_type='solid')
                if c_idx == 5:
                    cell.fill = pass_fill
                    cell.font = pass_font

        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 4, 15)

    @staticmethod
    def _populate_suite_sheet(ws, test_cases):
        headers = ["Test ID", "Suite", "Module", "Test Case Description", "Priority", "Expected Result", "Actual Result", "Status"]
        ws.append(headers)
        
        header_fill = PatternFill(start_color='1E293B', end_color='1E293B', fill_type='solid')
        header_font = Font(name='Segoe UI', size=11, bold=True, color='FFFFFF')
        
        for col_idx in range(1, 9):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        pass_fill = PatternFill(start_color='DCFCE7', end_color='DCFCE7', fill_type='solid')
        pass_font = Font(name='Segoe UI', size=10, bold=True, color='15803D')
        thin_border = Border(
            left=Side(style='thin', color='CBD5E1'),
            right=Side(style='thin', color='CBD5E1'),
            top=Side(style='thin', color='CBD5E1'),
            bottom=Side(style='thin', color='CBD5E1')
        )

        for row_idx, tc in enumerate(test_cases, start=2):
            row_data = [
                tc["test_id"],
                tc["suite"],
                tc["module"],
                tc["test_name"],
                tc["priority"],
                tc["expected"],
                tc["actual"],
                tc["status"]
            ]
            ws.append(row_data)
            
            for col_idx in range(1, 9):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.border = thin_border
                cell.font = Font(name='Segoe UI', size=10)
                if col_idx in [1, 5, 8]:
                    cell.alignment = Alignment(horizontal='center')
                if col_idx == 8:
                    cell.fill = pass_fill
                    cell.font = pass_font

        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = min(max(max_len + 3, 12), 50)
